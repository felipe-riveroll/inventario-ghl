"""
Módulo para generar reportes de Excel con openpyxl
"""
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment


class ExcelGenerator:
    """Generador de reportes de Excel para inventario"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_report(self, inventory_data: List[Dict], output_path: Optional[str] = None, progress_callback=None) -> str:
        """
        Crea un reporte de Excel con los datos del inventario
        
        Args:
            inventory_data: Lista de items del inventario
            output_path: Ruta donde guardar el archivo (opcional)
            progress_callback: Función callback para reportar progreso
            
        Returns:
            Ruta del archivo generado
        """
        if progress_callback:
            progress_callback("Creando estructura del reporte...")
        
        # Crear nuevo workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
        # Configurar hoja
        today = datetime.now().strftime("%d/%m/%Y")
        self.worksheet.title = f"Inventario_{today.replace('/', '-')}"
        
        # Crear encabezados
        self._create_headers()
        
        if progress_callback:
            progress_callback("Agregando datos del inventario...")
        
        # Agregar datos (incluye descarga de imágenes)
        self._add_data(inventory_data, progress_callback)
        
        if progress_callback:
            progress_callback("Aplicando formato...")
        
        # Aplicar formato
        self._apply_formatting()
        
        # Ajustar ancho de columnas
        self._adjust_column_widths()
        
        # Generar nombre de archivo si no se proporciona
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"inventario_ghl_{timestamp}.xlsx"
        
        if progress_callback:
            progress_callback("Guardando archivo...")
        
        # Guardar archivo
        self.workbook.save(output_path)
        
        return output_path
    
    def _create_headers(self):
        """Crea los encabezados de la tabla"""
        headers = ['Nombre', 'Nombre de producto', 'Cantidad disponible', 'Imagen']
        
        # Configurar altura del encabezado
        self.worksheet.row_dimensions[1].height = 30
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Agregar bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    def _add_data(self, inventory_data: List[Dict], progress_callback=None):
        """Agrega los datos del inventario a la hoja"""
        total_items = len(inventory_data)
        
        for i, item in enumerate(inventory_data):
            row = i + 2  # Empezar en fila 2
            
            if progress_callback and i % 10 == 0:  # Actualizar cada 10 productos
                progress = f"Procesando producto {i + 1} de {total_items}"
                progress_callback(progress)
            
            # Nombre
            self.worksheet.cell(row=row, column=1, value=item.get('Nombre', ''))
            
            # Nombre de producto
            self.worksheet.cell(row=row, column=2, value=item.get('Nombre de producto', ''))
            
            # Cantidad disponible
            cantidad = item.get('Cantidad disponible', 0)
            cell_cantidad = self.worksheet.cell(row=row, column=3, value=cantidad)
            cell_cantidad.alignment = Alignment(horizontal="center")
            
            # Configurar altura de fila a 100px (aproximadamente 75 puntos)
            self.worksheet.row_dimensions[row].height = 75
            
            # Imagen del producto - última solución: escribir como texto para conversión manual
            imagen_url = item.get('Imagen', '')
            if imagen_url and imagen_url.strip():
                cell_imagen = self.worksheet.cell(row=row, column=4)
                # openpyxl siempre añade @ a las fórmulas, escribir como texto
                formula_completa = f'=IMAGEN("{imagen_url}",1)'
                cell_imagen.value = formula_completa
                cell_imagen.data_type = 's'  # Tipo string para evitar procesamiento
                
                # Agregar comentario con instrucciones
                comment = Comment(
                    'Para activar la imagen:\n1. Haz clic en la celda\n2. Presiona F2 para editar\n3. Presiona Enter sin cambiar nada',
                    author='Sistema'
                )
                cell_imagen.comment = comment
            else:
                self.worksheet.cell(row=row, column=4, value='Sin imagen')
    
    def _apply_formatting(self):
        """Aplica formato general a la tabla"""
        # Bordes para todas las celdas con datos
        max_row = self.worksheet.max_row
        max_col = self.worksheet.max_column
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                self.worksheet.cell(row=row, column=col).border = thin_border
        
        # Alternar colores de fila
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row in range(2, max_row + 1):
            if row % 2 == 0:  # Filas pares
                for col in range(1, max_col + 1):
                    self.worksheet.cell(row=row, column=col).fill = light_fill
    
    def _adjust_column_widths(self):
        """Ajusta el ancho de las columnas automáticamente"""
        column_widths = {
            'A': 20,  # Nombre
            'B': 25,  # Nombre de producto
            'C': 18,  # Cantidad disponible
            'D': 20   # Imagen con fórmula IMAGEN()
        }
        
        for column, width in column_widths.items():
            self.worksheet.column_dimensions[column].width = width
    
    def add_summary(self, inventory_data: List[Dict]):
        """
        Agrega un resumen al final del reporte
        
        Args:
            inventory_data: Lista de items del inventario
        """
        if not self.worksheet:
            return
        
        # Encontrar la última fila con datos
        last_row = self.worksheet.max_row
        summary_row = last_row + 3
        
        # Total de productos
        total_productos = len(inventory_data)
        self.worksheet.cell(row=summary_row, column=1, value="Total de productos:")
        cell_total = self.worksheet.cell(row=summary_row, column=2, value=total_productos)
        cell_total.font = Font(bold=True)
        
        # Total de cantidad disponible
        total_cantidad = sum(item.get('Cantidad disponible', 0) for item in inventory_data)
        self.worksheet.cell(row=summary_row + 1, column=1, value="Total cantidad disponible:")
        cell_cantidad = self.worksheet.cell(row=summary_row + 1, column=2, value=total_cantidad)
        cell_cantidad.font = Font(bold=True)
        
        # Fecha de generación
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.worksheet.cell(row=summary_row + 2, column=1, value="Fecha de generación:")
        cell_fecha = self.worksheet.cell(row=summary_row + 2, column=2, value=fecha_generacion)
        cell_fecha.font = Font(bold=True)